# -*- coding: utf-8 -*-
"""
MetaERP 工单三维度 LLM 分类 v3

设计要点（对应 SKILL.md 的架构约定）：
- 判据唯一事实源是 references/spec.md，本脚本运行时读取全文拼装 system prompt，不硬编码判据
- 单条工单一次调用、独立上下文、temperature=0、线程池并发 —— 无批内交叉污染，结果可复现
- results.jsonl 逐条断点续跑；失败条目带 error 字段，--retry-errors 补跑
- 输出 JSON 先"依据"后结论（schema 字段顺序即推理顺序），并带"边界"标记驱动定向人工复核
- 枚举校验：LLM 输出不在词表内 → 记为 error，不写入脏值

子命令：
  ping                      连通性诊断（网关可达 + key 有效 + 模型能按 schema 输出）
  run --input ... --results-dir ... --output ...     全量分类（断点续跑）+ 写回
  finalize --input ... --results-dir ... --output ...  仅写回（不调 LLM）

用法示例见 SKILL.md。在 PowerShell 中直接 python scripts/classify.py ...，无内联代码。
"""
import argparse
import json
import os
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import httpx
import pandas as pd
import yaml
from anthropic import Anthropic

SKILL_ROOT = Path(__file__).resolve().parent.parent
SPEC_PATH = SKILL_ROOT / "references" / "spec.md"
CONFIG_PATH = Path(__file__).resolve().parent / "config.yaml"

# ---------------- 输出词表（与 spec.md 第〇章 0.2 保持一致） ----------------
TOPICS = ["外部/颗粒", "IT运维/基础设施", "IT开发/测试问题", "IT缺陷/性能问题",
          "IT需求", "权限专题", "其他业务"]
IT_TOPICS = set(TOPICS[:5])  # 命中即停：不判 L1/L2/解决方案
L1S = ["1-找不到页面入口/功能", "2-系统使用咨询", "3-操作失败/操作不生效",
       "4-业务数据异常/原因定位", "5-帮助查询/确认数据", "6-帮助维护配置/操作",
       "7-查不到数据", "8-系统逻辑/规则咨询"]
L2S = ["2a-操作步骤指导", "2b-如何申请权限", "3a-异常报错", "3b-功能入口不可用",
       "3c-选值选不到/选不了", "8a-规则/逻辑咨询", "8b-功能能力咨询",
       "8c-字段/接口/数据模型咨询"]
SOLUTIONS = ["S1-申请功能权限（角色）", "S2-申请数据权限（数据维度）", "S3-操作指引",
             "S4a-澄清业务规则", "S4b-诊断数据异常", "S4c-解释权限规则",
             "S5-数据查询", "S6-后台配置/修复", "S7-人工介入"]
L2_PARENT = {"2a": "2", "2b": "2", "3a": "3", "3b": "3", "3c": "3",
             "8a": "8", "8b": "8", "8c": "8"}

FIXED_HEADER = """你是 MetaERP 用户事件工单的分类专家。对给定的一条工单，按下方《判定规范》从三个维度分类：专题、问题表象 L1/L2、解决方案分类。

输出要求：只输出一个 JSON 对象，不加任何前后文字、不加 Markdown 代码围栏。字段按以下顺序（先陈述依据再给结论）：

{
  "依据": "<一句话：引用规范中命中的关键判据，说明归类理由>",
  "边界": <true|false，按规范第〇章 0.3 的边界标记规则>,
  "专题": "<枚举值>",
  "L1": "<枚举值或 null>",
  "L2": "<枚举值或 null>",
  "解决方案": "<枚举值或 null>"
}

硬性规则：
1. 所有类别名必须逐字取自规范第〇章 0.2 的输出词表，禁止自创、缩写或改写。
2. 专题命中 IT/外部类五类之一时，L1、L2、解决方案必须为 null。
3. L1 不属于 2/3/8 类时，L2 必须为 null。
4. L2 必须与 L1 匹配（2a/2b 属于 2 类，3a/3b/3c 属于 3 类，8a/8b/8c 属于 8 类）。
5. 严格按规范的判定顺序执行；判据冲突时按规范中的优先级规则裁定，并将"边界"置 true。

以下是《判定规范》全文：

"""


# ---------------- 基础设施 ----------------
def load_env():
    """从 skill 根目录 .env 自动加载凭证（不覆盖已有环境变量）"""
    env_file = SKILL_ROOT / ".env"
    if env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())


def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f)


def build_client(cfg):
    base_url = os.environ.get("LLM_BASE_URL")
    token = os.environ.get("ANTHROPIC_AUTH_TOKEN")
    if not base_url or not token:
        sys.exit("缺少 LLM_BASE_URL / ANTHROPIC_AUTH_TOKEN，请配置 .env（参考 .env.example）")
    # 内网直连：trust_env=False 绕过系统代理
    http_client = httpx.Client(trust_env=False, timeout=cfg["llm"]["timeout"])
    return Anthropic(base_url=base_url, auth_token=token, http_client=http_client)


def build_system_prompt() -> str:
    spec = SPEC_PATH.read_text(encoding="utf-8")
    return FIXED_HEADER + spec


def resolve_columns(df: pd.DataFrame, cfg) -> dict:
    """按 config 候选名探测列；探测失败即中止并提示配置 name_override"""
    resolved = {}
    for key, spec in cfg["excel"]["columns"].items():
        name = spec.get("name_override")
        if not name:
            name = next((c for c in spec["candidates"] if c in df.columns), None)
        if not name or name not in df.columns:
            sys.exit(f"未找到字段 [{key}]，候选 {spec['candidates']} 均不在表头中。"
                     f"实际表头：{list(df.columns)}。请在 config.yaml 的 "
                     f"excel.columns.{key}.name_override 显式指定。")
        resolved[key] = name
    return resolved


# ---------------- LLM 调用与校验 ----------------
def call_llm(client, cfg, system_prompt: str, ticket: dict) -> str:
    user_msg = (f"【标题】{ticket['title']}\n"
                f"【问题描述】{ticket['description'] or '（空）'}\n"
                f"【解决方案】{ticket['solution'] or '（空）'}")
    kwargs = dict(model=os.environ["LLM_MODEL"],
                  max_tokens=cfg["llm"]["max_tokens"],
                  temperature=cfg["llm"]["temperature"],
                  system=system_prompt,
                  messages=[{"role": "user", "content": user_msg}])
    if cfg["llm"].get("stream", True):
        parts = []
        with client.messages.stream(**kwargs) as s:
            for text in s.text_stream:
                parts.append(text)
        return "".join(parts)
    resp = client.messages.create(**kwargs)
    return "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")


def parse_and_validate(raw: str) -> dict:
    """解析 LLM 输出并做枚举/一致性校验；不合法直接抛 ValueError"""
    text = raw.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.startswith("json"):
            text = text[4:]
    start, end = text.find("{"), text.rfind("}")
    if start < 0 or end < 0:
        raise ValueError(f"输出中未找到 JSON：{raw[:200]}")
    obj = json.loads(text[start:end + 1])

    topic = obj.get("专题")
    if topic not in TOPICS:
        raise ValueError(f"专题不在词表：{topic!r}")
    l1, l2, sol = obj.get("L1"), obj.get("L2"), obj.get("解决方案")

    if topic in IT_TOPICS:
        if not (l1 is None and l2 is None and sol is None):
            raise ValueError(f"IT/外部类专题 [{topic}] 的 L1/L2/解决方案必须为 null")
    else:
        if l1 not in L1S:
            raise ValueError(f"L1 不在词表：{l1!r}")
        if sol not in SOLUTIONS:
            raise ValueError(f"解决方案不在词表：{sol!r}")
        l1_num = l1.split("-")[0]
        if l1_num in ("2", "3", "8"):
            if l2 not in L2S:
                raise ValueError(f"L1={l1_num} 类必须有合法 L2，实际：{l2!r}")
            if L2_PARENT[l2.split("-")[0]] != l1_num:
                raise ValueError(f"L2 [{l2}] 与 L1 [{l1}] 不匹配")
        elif l2 is not None:
            raise ValueError(f"L1={l1_num} 类的 L2 必须为 null，实际：{l2!r}")

    if not isinstance(obj.get("边界"), bool):
        raise ValueError(f"边界字段必须为布尔值：{obj.get('边界')!r}")
    if not obj.get("依据"):
        raise ValueError("依据字段为空")
    return {"专题": topic, "L1": l1, "L2": l2, "解决方案": sol,
            "依据": str(obj["依据"]), "边界": obj["边界"]}


def classify_one(client, cfg, system_prompt, ticket, max_retries):
    last_err = None
    for attempt in range(max_retries):
        try:
            raw = call_llm(client, cfg, system_prompt, ticket)
            result = parse_and_validate(raw)
            return {"ticket_id": ticket["ticket_id"], **result}
        except Exception as e:  # 网关波动 / JSON 坏 / 枚举校验失败，统一重试
            last_err = e
            time.sleep(2 ** attempt)
    return {"ticket_id": ticket["ticket_id"], "error": f"{type(last_err).__name__}: {last_err}"}


# ---------------- 子命令 ----------------
def cmd_ping(cfg):
    load_env()
    client = build_client(cfg)
    system_prompt = build_system_prompt()
    print(f"spec.md 已加载：{len(system_prompt)} 字符（含固定头）")
    demo = {"ticket_id": "PING", "title": "申请PO Query权限",
            "description": "查询采购订单需要什么权限，麻烦告知申请入口",
            "solution": "指导用户在OnePro申请PO Query角色权限"}
    r = classify_one(client, cfg, system_prompt, demo, max_retries=1)
    print(json.dumps(r, ensure_ascii=False, indent=2))
    if "error" in r:
        sys.exit("ping 失败：检查 .env 凭证 / 网关可达性 / 模型对中文 JSON schema 的遵循")
    print("ping OK：网关可达、key 有效、schema 遵循正常")


def read_tickets(input_path, cfg):
    df = pd.read_excel(input_path, dtype=str).fillna("")
    cols = resolve_columns(df, cfg)
    tickets = [{"ticket_id": row[cols["ticket_id"]] or f"ROW{i}",
                "title": row[cols["title"]],
                "description": row[cols["description"]],
                "solution": row[cols["solution"]]}
               for i, row in df.iterrows()]
    return df, cols, tickets


def cmd_run(args, cfg):
    load_env()
    client = build_client(cfg)
    system_prompt = build_system_prompt()
    results_dir = Path(args.results_dir)
    results_dir.mkdir(parents=True, exist_ok=True)
    results_path = results_dir / cfg["run"]["results_file"]

    _, _, tickets = read_tickets(args.input, cfg)

    done, errored = {}, {}
    if results_path.exists():
        for line in results_path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            rec = json.loads(line)
            (errored if "error" in rec else done)[rec["ticket_id"]] = rec

    todo = [t for t in tickets if t["ticket_id"] not in done]
    if not args.retry_errors:
        todo = [t for t in todo if t["ticket_id"] not in errored]
    print(f"总 {len(tickets)} 条 | 已完成 {len(done)} | 历史失败 {len(errored)}"
          f"{'（本次补跑）' if args.retry_errors else '（跳过，--retry-errors 补跑）'}"
          f" | 本次待跑 {len(todo)}")

    lock = threading.Lock()
    counter = {"ok": 0, "err": 0}
    with open(results_path, "a", encoding="utf-8") as fout, \
            ThreadPoolExecutor(max_workers=cfg["run"]["concurrency"]) as pool:
        futures = {pool.submit(classify_one, client, cfg, system_prompt, t,
                               cfg["run"]["max_retries"]): t for t in todo}
        for i, fut in enumerate(as_completed(futures), 1):
            rec = fut.result()
            with lock:
                fout.write(json.dumps(rec, ensure_ascii=False) + "\n")
                fout.flush()
                counter["err" if "error" in rec else "ok"] += 1
            if i % 50 == 0 or i == len(todo):
                print(f"  进度 {i}/{len(todo)}  成功 {counter['ok']}  失败 {counter['err']}")

    if counter["err"]:
        print(f"存在 {counter['err']} 条失败，稳定后用 --retry-errors 补跑，再 finalize")
    finalize(args.input, results_path, args.output, cfg)


def finalize(input_path, results_path, output_path, cfg):
    """写回 Excel：openpyxl 保留原格式，新增分类列并复制表头样式"""
    from openpyxl import load_workbook
    from copy import copy

    results = {}
    for line in Path(results_path).read_text(encoding="utf-8").splitlines():
        if line.strip():
            rec = json.loads(line)
            if "error" not in rec:
                results[rec["ticket_id"]] = rec

    df = pd.read_excel(input_path, dtype=str).fillna("")
    cols = resolve_columns(df, cfg)
    wb = load_workbook(input_path)
    ws = wb.active
    header_row = 1
    id_col_idx = next(c.column for c in ws[header_row] if c.value == cols["ticket_id"])
    style_src = ws.cell(row=header_row, column=1)

    out = cfg["excel"]["out_columns"]
    new_cols = [(out["topic"], "专题"), (out["l1"], "L1"), (out["l2"], "L2"),
                (out["solution_class"], "解决方案"), (out["rationale"], "依据"),
                (out["boundary"], "边界")]
    base = ws.max_column
    for j, (col_name, _) in enumerate(new_cols, 1):
        cell = ws.cell(row=header_row, column=base + j, value=col_name)
        cell.font, cell.fill = copy(style_src.font), copy(style_src.fill)
        cell.alignment, cell.border = copy(style_src.alignment), copy(style_src.border)

    matched = 0
    for r in range(header_row + 1, ws.max_row + 1):
        tid = ws.cell(row=r, column=id_col_idx).value
        rec = results.get(str(tid) if tid is not None else f"ROW{r - header_row - 1}")
        if not rec:
            continue
        matched += 1
        for j, (_, key) in enumerate(new_cols, 1):
            v = rec.get(key)
            if key == "边界":
                v = "是" if v else ""
            ws.cell(row=r, column=base + j, value="" if v is None else v)

    out_path = Path(output_path)
    try:
        wb.save(out_path)
    except PermissionError:  # 文件被 Excel 占用
        out_path = out_path.with_name(out_path.stem + "_分类结果" + out_path.suffix)
        wb.save(out_path)
    print(f"写回完成：{out_path}（匹配 {matched}/{len(results)} 条结果）")
    boundary_n = sum(1 for r in results.values() if r.get("边界"))
    print(f"边界工单 {boundary_n} 条 —— 人工复核时按 [{out['boundary']}]=是 筛选定向审核")


def main():
    p = argparse.ArgumentParser(description="MetaERP 工单 LLM 分类 v3")
    sub = p.add_subparsers(dest="cmd", required=True)
    sub.add_parser("ping")
    for name in ("run", "finalize"):
        sp = sub.add_parser(name)
        sp.add_argument("--input", required=True)
        sp.add_argument("--results-dir", required=True)
        sp.add_argument("--output", required=True)
        if name == "run":
            sp.add_argument("--retry-errors", action="store_true")
    args = p.parse_args()
    cfg = load_config()
    if args.cmd == "ping":
        cmd_ping(cfg)
    elif args.cmd == "run":
        cmd_run(args, cfg)
    else:
        load_env()
        finalize(args.input, Path(args.results_dir) / cfg["run"]["results_file"],
                 args.output, cfg)


if __name__ == "__main__":
    main()
